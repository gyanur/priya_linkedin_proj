from django.shortcuts import render , HttpResponse
import requests 
from bs4 import BeautifulSoup 
import openpyxl
from openpyxl.styles import Font


# Create your views here.
def index(request):
    return render(request,"index.html")


def excel(request):
    return render(request,"excel.html")


def get_linkedin_link(company_name):
    search_query = f'site:linkedin.com {company_name}'
    url = f'https://www.google.com/search?q={search_query}'

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36'
    }

    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, 'html.parser')

    # Extract the first search result link
    link_element = soup.find('div', class_='yuRUbf')
    if link_element:
        link = link_element.a['href']
        return link
    return None

def search1(request):
    if request.method == 'POST':
        company_name = request.POST['company_name']
        linkedin_link = get_linkedin_link(company_name)
        return render(request, 'index.html', {'linkedin_link': linkedin_link})
    return render(request, 'index.html')
        
        
def upload_excel_ceo(request):
    if "GET" == request.method:
        return render(request, 'excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)
        
        # Add a new column header for the LinkedIn links
        header_cell = worksheet.cell(row=1, column=worksheet.max_column + 1)
        header_cell.value = "Company CEO Linkedin Link"
        header_cell.font = Font(bold=True)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows(min_row=2):
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))

                search_query = f'site:linkedin.com {cell.value} CEO'
                url = f'https://www.google.com/search?q={search_query}'

                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36'
                }

                response = requests.get(url, headers=headers)
                soup = BeautifulSoup(response.text, 'html.parser')

                # Extract the first search result link
                link_element = soup.find('div', class_='yuRUbf')
                if link_element:
                    link = link_element.a['href']
                    print(link)
                    excel_data.append(row_data)
                    
                    link_cell = worksheet.cell(row=row[0].row, column=worksheet.max_column)
                    link_cell.value = link

        # Save the modified workbook
        modified_excel_file = "ceo_list.xlsx"
        wb.save(modified_excel_file)

        # Create a response with the modified Excel file
        with open(modified_excel_file, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="ceo_list.xlsx"'
                    
                # return render(request,"index.html",{"link":link})


        return render(request, 'excel.html', {"excel_data":excel_data})
    
    


def upload_excel_cto(request):
    if "GET" == request.method:
        return render(request, 'excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)
        
        # Add a new column header for the LinkedIn links
        header_cell = worksheet.cell(row=1, column=worksheet.max_column + 1)
        header_cell.value = "Company CTO Linkedin Link"
        header_cell.font = Font(bold=True)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows(min_row=2):
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))

                search_query = f'site:linkedin.com {cell.value} CTO'
                url = f'https://www.google.com/search?q={search_query}'

                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36'
                }

                response = requests.get(url, headers=headers)
                soup = BeautifulSoup(response.text, 'html.parser')

                # Extract the first search result link
                link_element = soup.find('div', class_='yuRUbf')
                if link_element:
                    link = link_element.a['href']
                    print(link)
                    excel_data.append(row_data)
                    
                    link_cell = worksheet.cell(row=row[0].row, column=worksheet.max_column)
                    link_cell.value = link

        # Save the modified workbook
        modified_excel_file = "cto_list.xlsx"
        wb.save(modified_excel_file)

        # Create a response with the modified Excel file
        with open(modified_excel_file, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="cto_list.xlsx"'
                    
                # return render(request,"index.html",{"link":link})


        return render(request, 'excel.html', {"excel_data":excel_data})
    
    
    

def upload_excel_founders(request):
    if "GET" == request.method:
        return render(request, 'excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)
        
        # Add a new column header for the LinkedIn links
        header_cell = worksheet.cell(row=1, column=worksheet.max_column + 1)
        header_cell.value = "Company Founders Linkedin Link"
        header_cell.font = Font(bold=True)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows(min_row=2):
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))

                search_query = f'site:linkedin.com {cell.value} Founders'
                url = f'https://www.google.com/search?q={search_query}'

                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36'
                }

                response = requests.get(url, headers=headers)
                soup = BeautifulSoup(response.text, 'html.parser')

                # Extract the first search result link
                link_element = soup.find('div', class_='yuRUbf')
                if link_element:
                    link = link_element.a['href']
                    print(link)
                    excel_data.append(row_data)
                    
                    link_cell = worksheet.cell(row=row[0].row, column=worksheet.max_column)
                    link_cell.value = link

        # Save the modified workbook
        modified_excel_file = "founders_list.xlsx"
        wb.save(modified_excel_file)

        # Create a response with the modified Excel file
        with open(modified_excel_file, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="founders_list.xlsx"'
                    
                # return render(request,"index.html",{"link":link})

        return render(request, 'excel.html', {"excel_data":excel_data})
    
    
    

def upload_excel_directors(request):
    if "GET" == request.method:
        return render(request, 'excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)
        
        # Add a new column header for the LinkedIn links
        header_cell = worksheet.cell(row=1, column=worksheet.max_column + 1)
        header_cell.value = "Company Directors Linkedin Link"
        header_cell.font = Font(bold=True)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows(min_row=2):
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))

                search_query = f'site:linkedin.com {cell.value} Directors'
                url = f'https://www.google.com/search?q={search_query}'

                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36'
                }

                response = requests.get(url, headers=headers)
                soup = BeautifulSoup(response.text, 'html.parser')

                # Extract the first search result link
                link_element = soup.find('div', class_='yuRUbf')
                if link_element:
                    link = link_element.a['href']
                    print(link)
                    excel_data.append(row_data)
                    
                    link_cell = worksheet.cell(row=row[0].row, column=worksheet.max_column)
                    link_cell.value = link

        # Save the modified workbook
        modified_excel_file = "Directors_list.xlsx"
        wb.save(modified_excel_file)

        # Create a response with the modified Excel file
        with open(modified_excel_file, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="Directors_list.xlsx"'
                    
                # return render(request,"index.html",{"link":link})

        return render(request, 'excel.html', {"excel_data":excel_data})